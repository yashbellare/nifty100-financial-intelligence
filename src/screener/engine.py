# ============================================================
# SPRINT 3 — DAY 17
# COMPOSITE QUALITY SCORE & EXCEL EXPORT
# ============================================================

from pathlib import Path
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SCREENER_OUTPUT_XLSX = OUTPUT_DIR / "screener_output.xlsx"

DEFAULT_DAY17_INPUT = OUTPUT_DIR / "nifty100_screened_results.csv"


# ============================================================
# DAY 17 — SIX PRESET SCREENERS
# ============================================================

DAY17_PRESETS = {

    "Quality Compounder": {
        "roe": (">", 15),
        "de": ("<", 1.0),
        "fcf": (">", 0),
        "revenue_cagr_5yr": (">", 10),
    },

    "Value Pick": {
        "pe": ("<", 20),
        "pb": ("<", 3.0),
        "de": ("<", 2.0),
        "dividend_yield": (">", 1),
    },

    "Growth Accelerator": {
        "pat_cagr_5yr": (">", 20),
        "revenue_cagr_5yr": (">", 15),
        "de": ("<", 2.0),
    },

    "Dividend Champion": {
        "dividend_yield": (">", 2),
        "dividend_payout": ("<", 80),
        "fcf": (">", 0),
    },

    "Debt-Free Blue Chip": {
        "de": ("=", 0),
        "roe": (">", 12),
        "revenue": (">", 5000),
    },

    "Turnaround Watch": {
        "revenue_cagr_3yr": (">", 10),
        "fcf": (">", 0),
        "de_declining": ("=", True),
    },
}


# ============================================================
# COLUMN ALIASES
# ============================================================

DAY17_ALIASES = {

    "company_id": [
        "company_id",
        "id",
        "companyid",
        "symbol",
        "ticker",
    ],

    "company_name": [
        "company_name",
        "name",
        "company",
        "companyname",
        "stock_name",
    ],

    "broad_sector": [
        "broad_sector",
        "sector",
        "sector_name",
        "industry",
        "Sector",
    ],

    # --------------------------------------------------------
    # Profitability
    # --------------------------------------------------------

    "roe": [
        "roe",
        "roe_percentage",
        "return_on_equity_pct",
        "return_on_equity",
        "return_on_equity_percentage",
        "ROE",
    ],

    "roce": [
        "roce",
        "roce_percentage",
        "return_on_capital_employed",
        "return_on_capital_employed_pct",
        "ROCE",
    ],

    "npm": [
        "npm",
        "net_profit_margin",
        "net_profit_margin_pct",
        "net_profit_margin_percentage",
        "NPM",
    ],

    # --------------------------------------------------------
    # Leverage
    # --------------------------------------------------------

    "de": [
        "de",
        "de_ratio",
        "debt_to_equity",
        "debt_to_equity_pct",
        "debt_to_equity_ratio",
        "D/E",
        "d_e",
        "debt_equity",
    ],

    "de_previous_year": [
        "de_previous_year",
        "de_prev_year",
        "previous_de",
        "de_last_year",
        "previous_year_de",
        "de_previous",
    ],

    # --------------------------------------------------------
    # Cash flow
    # --------------------------------------------------------

    "fcf": [
        "fcf",
        "free_cash_flow",
        "free_cash_flow_cr",
        "FCF",
    ],

    "cfo": [
        "cfo",
        "cash_from_operations",
        "cash_from_operations_cr",
        "cash_flow_from_operations",
    ],

    "fcf_cagr": [
        "fcf_cagr",
        "fcf_cagr_5yr",
        "FCF CAGR",
        "FCF_CAGR_5yr",
        "free_cash_flow_cagr",
    ],

    "cfo_pat": [
        "cfo_pat",
        "cfo_pat_ratio",
        "cfo_to_pat",
        "CFO/PAT",
        "cfo_pat_percentage",
    ],

    # --------------------------------------------------------
    # Growth
    # --------------------------------------------------------

    "revenue_cagr_3yr": [
        "revenue_cagr_3yr",
        "revenue_cagr_3yr_pct",
        "sales_cagr_3yr",
        "compounded_sales_growth_3yr",
    ],

    "revenue_cagr_5yr": [
        "revenue_cagr_5yr",
        "revenue_cagr_5yr_pct",
        "revenue_5yr_cagr",
        "sales_cagr_5yr",
        "compounded_sales_growth",
        "compounded_sales_growth_pct",
        "Revenue CAGR 5yr",
    ],

    "pat_cagr_5yr": [
        "pat_cagr_5yr",
        "pat_cagr_5yr_pct",
        "pat_5yr_cagr",
        "profit_cagr_5yr",
        "profit_growth_5yr",
        "compounded_profit_growth",
        "compounded_profit_growth_pct",
        "PAT CAGR 5yr",
    ],

    "eps_cagr_5yr": [
        "eps_cagr_5yr",
        "eps_cagr_5yr_pct",
        "eps_growth_5yr",
        "EPS CAGR 5yr",
    ],

    # --------------------------------------------------------
    # Valuation
    # --------------------------------------------------------

    "pe": [
        "pe",
        "pe_ratio",
        "price_to_earnings",
        "price_earnings",
        "P/E",
    ],

    "pb": [
        "pb",
        "pb_ratio",
        "price_to_book",
        "P/B",
    ],

    # --------------------------------------------------------
    # Dividend
    # --------------------------------------------------------

    "dividend_yield": [
        "dividend_yield",
        "dividend_yield_pct",
        "Dividend Yield",
    ],

    "dividend_payout": [
        "dividend_payout",
        "dividend_payout_ratio",
        "dividend_payout_ratio_pct",
        "dividend_payout_pct",
        "dividend_payout_percentage",
        "Dividend Payout",
    ],

    # --------------------------------------------------------
    # Other KPIs
    # --------------------------------------------------------

    "icr": [
        "icr",
        "interest_coverage",
        "interest_coverage_ratio",
        "ICR",
    ],

    "icr_label": [
        "icr_label",
        "interest_coverage_label",
    ],

    "market_cap": [
        "market_cap",
        "market_cap_cr",
        "market_cap_crore",
        "market_capitalization",
        "Market Cap",
    ],

    "net_profit": [
        "net_profit",
        "net_profit_cr",
        "Net Profit",
    ],

    "eps": [
        "eps",
        "earnings_per_share",
        "EPS",
    ],

    "asset_turnover": [
        "asset_turnover",
        "asset_turnover_ratio",
        "Asset Turnover",
    ],

    "sales": [
        "sales",
        "sales_cr",
        "revenue",
        "revenue_cr",
        "total_sales",
        "Revenue",
        "Sales",
    ],

    "revenue": [
        "revenue",
        "revenue_cr",
        "sales",
        "sales_cr",
        "total_sales",
        "Revenue",
        "Sales",
    ],

    "de_declining": [
        "de_declining",
        "de_declining_yoy",
        "D/E declining",
    ],

    "year": [
        "year",
        "financial_year",
        "fy",
        "date",
        "period",
    ],
}


# ============================================================
# COLUMN RESOLVER
# ============================================================

def _normalise_column_name(value):
    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
        .replace("%", "pct")
        .replace("(", "")
        .replace(")", "")
    )


def _resolve_column(df, logical_name):
    """
    Resolve a logical Day-17 metric to an actual DataFrame column.
    """

    if logical_name not in DAY17_ALIASES:
        return None

    exact_columns = {
        str(column).strip().lower(): column
        for column in df.columns
    }

    normalised_columns = {
        _normalise_column_name(column): column
        for column in df.columns
    }

    for alias in DAY17_ALIASES[logical_name]:

        if alias in df.columns:
            return alias

        alias_lower = str(alias).strip().lower()

        if alias_lower in exact_columns:
            return exact_columns[alias_lower]

        alias_normalised = _normalise_column_name(alias)

        if alias_normalised in normalised_columns:
            return normalised_columns[alias_normalised]

    return None


# ============================================================
# NUMERIC SERIES HELPER
# ============================================================

def _series(df, logical_name, default=np.nan):

    column = _resolve_column(df, logical_name)

    if column is None:
        return pd.Series(
            default,
            index=df.index,
            dtype="float64"
        )

    series = df[column]

    # Handle strings such as:
    # "18.5%"
    # "1,250.50"
    # "₹ 1,250"
    if series.dtype == object:

        series = (
            series.astype(str)
            .str.replace(",", "", regex=False)
            .str.replace("%", "", regex=False)
            .str.replace("₹", "", regex=False)
            .str.strip()
        )

    return pd.to_numeric(
        series,
        errors="coerce"
    )


# ============================================================
# BOOLEAN HELPER
# ============================================================

def _boolean_series(df, logical_name):

    column = _resolve_column(df, logical_name)

    if column is None:
        return pd.Series(
            False,
            index=df.index,
            dtype=bool
        )

    values = df[column]

    if values.dtype == bool:
        return values.fillna(False)

    return (
        values.astype(str)
        .str.strip()
        .str.lower()
        .isin([
            "true",
            "1",
            "yes",
            "y",
            "decreasing",
            "declining",
        ])
    )


# ============================================================
# P10 / P90 WINSORISATION + 0–100 NORMALISATION
# ============================================================

def _winsor_score(series):

    s = pd.to_numeric(
        series,
        errors="coerce"
    )

    valid = s.dropna()

    if valid.empty:
        return pd.Series(
            50.0,
            index=series.index,
            dtype=float
        )

    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)

    if pd.isna(p10) or pd.isna(p90):
        return pd.Series(
            50.0,
            index=series.index,
            dtype=float
        )

    if p10 == p90:
        return pd.Series(
            50.0,
            index=series.index,
            dtype=float
        )

    clipped = s.clip(
        lower=p10,
        upper=p90
    )

    score = (
        (clipped - p10)
        / (p90 - p10)
    ) * 100

    return score.fillna(50.0)


# ============================================================
# SECTOR-RELATIVE SCORE
# ============================================================

def _sector_relative_score(
    df,
    metric,
    sector_column
):

    values = _series(
        df,
        metric
    )

    if sector_column is None:

        return _winsor_score(
            values
        )

    result = pd.Series(
        50.0,
        index=df.index,
        dtype=float
    )

    sectors = df[sector_column]

    grouped = values.groupby(
        sectors,
        dropna=False
    )

    for _, index_values in grouped:

        score = _winsor_score(
            values.loc[index_values.index]
        )

        result.loc[index_values.index] = score

    return result


# ============================================================
# INVERSE SCORE
# LOWER VALUE = BETTER
# ============================================================

def _inverse_sector_score(
    df,
    metric,
    sector_column
):

    normal_score = _sector_relative_score(
        df,
        metric,
        sector_column
    )

    return (
        100 - normal_score
    ).clip(0, 100)


# ============================================================
# CFO / PAT
# ============================================================

def _get_cfo_pat(df):

    direct = _resolve_column(
        df,
        "cfo_pat"
    )

    if direct is not None:

        return _series(
            df,
            "cfo_pat"
        )

    cfo = _series(
        df,
        "cfo"
    )

    pat = _series(
        df,
        "net_profit"
    )

    result = pd.Series(
        np.nan,
        index=df.index,
        dtype=float
    )

    valid = (
        pat.notna()
        & cfo.notna()
        & pat.ne(0)
    )

    result.loc[valid] = (
        cfo.loc[valid]
        / pat.loc[valid]
    )

    return result


# ============================================================
# FCF CAGR
# ============================================================

def _get_fcf_cagr(df):

    direct = _resolve_column(
        df,
        "fcf_cagr"
    )

    if direct is not None:

        return _series(
            df,
            "fcf_cagr"
        )

    # If the project already has FCF CAGR, use it.
    # Otherwise keep neutral score rather than inventing
    # historical data from a single-year record.

    return pd.Series(
        np.nan,
        index=df.index,
        dtype=float
    )


# ============================================================
# ICR SCORE
# DEBT-FREE = INFINITY = 100
# ============================================================

def _icr_score(
    df,
    sector_column
):

    icr = _series(
        df,
        "icr"
    )

    de = _series(
        df,
        "de"
    )

    debt_free = de.eq(0)

    # Explicitly handle Debt Free.
    icr_adjusted = icr.copy()

    icr_adjusted.loc[debt_free] = np.inf

    finite_icr = icr_adjusted.replace(
        [np.inf, -np.inf],
        np.nan
    )

    score = _sector_relative_score_from_series(
        finite_icr,
        df,
        sector_column
    )

    score.loc[debt_free] = 100.0

    return score.clip(
        0,
        100
    )


# ============================================================
# GENERIC SECTOR SCORE FROM SERIES
# ============================================================

def _sector_relative_score_from_series(
    values,
    df,
    sector_column
):

    if sector_column is None:
        return _winsor_score(
            values
        )

    result = pd.Series(
        50.0,
        index=df.index,
        dtype=float
    )

    grouped = values.groupby(
        df[sector_column],
        dropna=False
    )

    for _, group in grouped:

        result.loc[group.index] = (
            _winsor_score(group)
        )

    return result


# ============================================================
# FCF POSITIVE SCORE
# ============================================================

def _fcf_positive_score(df):

    fcf = _series(
        df,
        "fcf"
    )

    result = pd.Series(
        0.0,
        index=df.index,
        dtype=float
    )

    result.loc[fcf > 0] = 100.0

    return result


# ============================================================
# DE DECLINING
# ============================================================

def _calculate_de_declining(df):

    direct = _resolve_column(
        df,
        "de_declining"
    )

    if direct is not None:

        return _boolean_series(
            df,
            "de_declining"
        )

    current_de = _series(
        df,
        "de"
    )

    previous_de = _series(
        df,
        "de_previous_year"
    )

    result = pd.Series(
        False,
        index=df.index,
        dtype=bool
    )

    valid = (
        current_de.notna()
        & previous_de.notna()
    )

    result.loc[valid] = (
        current_de.loc[valid]
        < previous_de.loc[valid]
    )

    return result


# ============================================================
# COMPOSITE QUALITY SCORE
# ============================================================

def calculate_composite_quality_score(
    df,
    sector_relative=True
):

    result = df.copy()

    sector_column = _resolve_column(
        result,
        "broad_sector"
    )

    # --------------------------------------------------------
    # Profitability — 35%
    # --------------------------------------------------------

    if sector_relative:

        roe_score = _sector_relative_score(
            result,
            "roe",
            sector_column
        )

        roce_score = _sector_relative_score(
            result,
            "roce",
            sector_column
        )

        npm_score = _sector_relative_score(
            result,
            "npm",
            sector_column
        )

    else:

        roe_score = _winsor_score(
            _series(result, "roe")
        )

        roce_score = _winsor_score(
            _series(result, "roce")
        )

        npm_score = _winsor_score(
            _series(result, "npm")
        )

    profitability = (
        roe_score * 0.15
        + roce_score * 0.10
        + npm_score * 0.10
    )

    # --------------------------------------------------------
    # Cash Quality — 30%
    # --------------------------------------------------------

    fcf_cagr = _get_fcf_cagr(
        result
    )

    cfo_pat = _get_cfo_pat(
        result
    )

    if sector_relative:

        fcf_cagr_score = (
            _sector_relative_score_from_series(
                fcf_cagr,
                result,
                sector_column
            )
        )

        cfo_pat_score = (
            _sector_relative_score_from_series(
                cfo_pat,
                result,
                sector_column
            )
        )

    else:

        fcf_cagr_score = _winsor_score(
            fcf_cagr
        )

        cfo_pat_score = _winsor_score(
            cfo_pat
        )

    fcf_positive_score = (
        _fcf_positive_score(result)
    )

    cash_quality = (
        fcf_cagr_score * 0.15
        + cfo_pat_score * 0.10
        + fcf_positive_score * 0.05
    )

    # --------------------------------------------------------
    # Growth — 20%
    # --------------------------------------------------------

    if sector_relative:

        revenue_growth_score = (
            _sector_relative_score(
                result,
                "revenue_cagr_5yr",
                sector_column
            )
        )

        pat_growth_score = (
            _sector_relative_score(
                result,
                "pat_cagr_5yr",
                sector_column
            )
        )

    else:

        revenue_growth_score = _winsor_score(
            _series(
                result,
                "revenue_cagr_5yr"
            )
        )

        pat_growth_score = _winsor_score(
            _series(
                result,
                "pat_cagr_5yr"
            )
        )

    growth = (
        revenue_growth_score * 0.10
        + pat_growth_score * 0.10
    )

    # --------------------------------------------------------
    # Leverage — 15%
    # --------------------------------------------------------

    if sector_relative:

        de_score = _inverse_sector_score(
            result,
            "de",
            sector_column
        )

    else:

        de_score = (
            100
            - _winsor_score(
                _series(result, "de")
            )
        )

    icr_score = _icr_score(
        result,
        sector_column
    )

    leverage = (
        de_score * 0.10
        + icr_score * 0.05
    )

    # --------------------------------------------------------
    # Final score
    # --------------------------------------------------------

    result["composite_quality_score"] = (
        profitability
        + cash_quality
        + growth
        + leverage
    )

    result["composite_quality_score"] = (
        result["composite_quality_score"]
        .clip(0, 100)
        .round(2)
    )

    # --------------------------------------------------------
    # Sort descending
    # --------------------------------------------------------

    result = result.sort_values(
        "composite_quality_score",
        ascending=False,
        kind="mergesort"
    ).reset_index(drop=True)

    return result


# ============================================================
# PRESET FILTER
# ============================================================

def _apply_day17_preset_check(
    df,
    preset_name
):

    result = df.copy()

    rules = DAY17_PRESETS[
        preset_name
    ]

    mask = pd.Series(
        True,
        index=result.index
    )

    # Create D/E declining if necessary.
    if "de_declining" in rules:

        result["_day17_de_declining"] = (
            _calculate_de_declining(result)
        )

    for metric, rule in rules.items():

        operator, threshold = rule

        # ----------------------------------------------------
        # Boolean rule
        # ----------------------------------------------------

        if metric == "de_declining":

            values = result[
                "_day17_de_declining"
            ]

            mask &= (
                values
                == bool(threshold)
            )

            continue

        # ----------------------------------------------------
        # Numeric rule
        # ----------------------------------------------------

        values = _series(
            result,
            metric
        )

        if operator == ">":

            mask &= values > threshold

        elif operator == ">=":

            mask &= values >= threshold

        elif operator == "<":

            mask &= values < threshold

        elif operator == "<=":

            mask &= values <= threshold

        elif operator == "=":

            mask &= np.isclose(
                values,
                threshold,
                equal_nan=False
            )

    result = result.loc[
        mask
    ].copy()

    result.drop(
        columns=["_day17_de_declining"],
        errors="ignore",
        inplace=True
    )

    return result


# ============================================================
# 20 KPI COLUMNS
# ============================================================

DAY17_KPI_KEYS = [

    "company_id",
    "company_name",
    "broad_sector",

    "roe",
    "roce",
    "npm",

    "de",
    "fcf",
    "fcf_cagr",
    "cfo_pat",

    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    "icr",

    "market_cap",
    "net_profit",
    "eps",
    "asset_turnover",

    "dividend_yield",

    "composite_quality_score",
]


# ============================================================
# PREPARE KPI OUTPUT
# ============================================================

def _prepare_kpi_output(df):

    output = pd.DataFrame(
        index=df.index
    )

    for key in DAY17_KPI_KEYS:

        if key == "composite_quality_score":

            output[
                "composite_quality_score"
            ] = df[
                "composite_quality_score"
            ]

            continue

        # CFO/PAT can be derived.
        if key == "cfo_pat":

            output["cfo_pat"] = (
                _get_cfo_pat(df)
            )

            continue

        # FCF CAGR can be derived/loaded.
        if key == "fcf_cagr":

            output["fcf_cagr"] = (
                _get_fcf_cagr(df)
            )

            continue

        column = _resolve_column(
            df,
            key
        )

        if column is not None:

            output[key] = df[
                column
            ]

        else:

            output[key] = np.nan

    return output


# ============================================================
# EXCEL THRESHOLD CHECK
# ============================================================

def _cell_meets_threshold(
    value,
    metric,
    preset_name
):

    if pd.isna(value):
        return False

    rules = DAY17_PRESETS.get(
        preset_name,
        {}
    )

    if metric not in rules:
        return False

    operator, threshold = (
        rules[metric]
    )

    # Boolean
    if metric == "de_declining":

        return (
            bool(value)
            == bool(threshold)
        )

    try:

        numeric_value = float(
            value
        )

    except (
        TypeError,
        ValueError
    ):

        return False

    if operator == ">":
        return numeric_value > threshold

    if operator == ">=":
        return numeric_value >= threshold

    if operator == "<":
        return numeric_value < threshold

    if operator == "<=":
        return numeric_value <= threshold

    if operator == "=":
        return np.isclose(
            numeric_value,
            threshold
        )

    return False


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_screener_output(
    preset_results,
    output_path=SCREENER_OUTPUT_XLSX
):

    output_path = Path(
        output_path
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb = Workbook()

    # Remove default sheet.
    default_sheet = wb.active
    wb.remove(default_sheet)

    # --------------------------------------------------------
    # Styles
    # --------------------------------------------------------

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    header_font = Font(
        bold=True
    )

    # --------------------------------------------------------
    # Create six sheets
    # --------------------------------------------------------

    for preset_name, df in (
        preset_results.items()
    ):

        sheet_name = preset_name[:31]

        ws = wb.create_sheet(
            title=sheet_name
        )

        kpi_df = _prepare_kpi_output(
            df
        )

        # ----------------------------------------------------
        # Header
        # ----------------------------------------------------

        for column_index, column in enumerate(
            kpi_df.columns,
            start=1
        ):

            cell = ws.cell(
                row=1,
                column=column_index,
                value=column
            )

            cell.fill = header_fill
            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # ----------------------------------------------------
        # Data
        # ----------------------------------------------------

        for row_index, (_, row) in enumerate(
            kpi_df.iterrows(),
            start=2
        ):

            for column_index, column in enumerate(
                kpi_df.columns,
                start=1
            ):

                value = row[column]

                if pd.isna(value):

                    excel_value = None

                else:

                    excel_value = value

                cell = ws.cell(
                    row=row_index,
                    column=column_index,
                    value=excel_value
                )

                # ------------------------------------------------
                # Preset threshold colour
                # ------------------------------------------------

                if column in DAY17_PRESETS.get(
                    preset_name,
                    {}
                ):

                    meets = (
                        _cell_meets_threshold(
                            value,
                            column,
                            preset_name
                        )
                    )

                    if meets:

                        cell.fill = green_fill

                    else:

                        cell.fill = red_fill

        # ----------------------------------------------------
        # Freeze header
        # ----------------------------------------------------

        ws.freeze_panes = "A2"

        # ----------------------------------------------------
        # Auto filter
        # ----------------------------------------------------

        if ws.max_row >= 1:

            ws.auto_filter.ref = (
                ws.dimensions
            )

        # ----------------------------------------------------
        # Header row height
        # ----------------------------------------------------

        ws.row_dimensions[
            1
        ].height = 24

        # ----------------------------------------------------
        # Auto width
        # ----------------------------------------------------

        for column_index in range(
            1,
            len(kpi_df.columns) + 1
        ):

            column_letter = (
                get_column_letter(
                    column_index
                )
            )

            max_length = 0

            for cell in ws[
                column_letter
            ]:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max(
                    max_length + 2,
                    12
                ),
                30
            )

        # ----------------------------------------------------
        # Number formatting
        # ----------------------------------------------------

        for row in ws.iter_rows(
            min_row=2
        ):

            for cell in row:

                if isinstance(
                    cell.value,
                    (int, float, np.integer, np.floating)
                ):

                    cell.number_format = (
                        "0.00"
                    )

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    wb.save(
        output_path
    )

    return output_path


# ============================================================
# LOAD DAY 17 DATA
# ============================================================

def _load_day17_data():

    # --------------------------------------------------------
    # 1. Existing loader functions
    # --------------------------------------------------------

    for loader_name in [
        "load_financial_ratios",
        "load_screener_data",
        "load_data",
    ]:

        loader = globals().get(
            loader_name
        )

        if callable(loader):

            try:

                loaded = loader()

                if isinstance(
                    loaded,
                    pd.DataFrame
                ):

                    return loaded

            except Exception:
                pass

    # --------------------------------------------------------
    # 2. Existing CSV produced by Day 15/16
    # --------------------------------------------------------

    if DEFAULT_DAY17_INPUT.exists():

        df = pd.read_csv(
            DEFAULT_DAY17_INPUT
        )

        if not df.empty:

            return df

    # --------------------------------------------------------
    # 3. Search other CSV candidates
    # --------------------------------------------------------

    candidates = [

        OUTPUT_DIR
        / "financial_ratios.csv",

        OUTPUT_DIR
        / "nifty100_screened_results.csv",

        PROJECT_ROOT
        / "data"
        / "financial_ratios.csv",

        PROJECT_ROOT
        / "data"
        / "financial_ratios.xlsx",

    ]

    for candidate in candidates:

        if not candidate.exists():
            continue

        try:

            if candidate.suffix.lower() == ".csv":

                df = pd.read_csv(
                    candidate
                )

            else:

                df = pd.read_excel(
                    candidate
                )

            if isinstance(
                df,
                pd.DataFrame
            ) and not df.empty:

                return df

        except Exception:
            continue

    raise RuntimeError(
        "Day 17 could not find the financial "
        "ratios DataFrame. Expected "
        f"{DEFAULT_DAY17_INPUT}. "
        "You can also call run_day17(df) "
        "with a pandas DataFrame."
    )


# ============================================================
# DAY 17 MAIN RUNNER
# ============================================================

def run_day17(df=None):

    print()
    print("=" * 70)
    print("SPRINT 3 — DAY 17")
    print("COMPOSITE SCORE & EXCEL EXPORT")
    print("=" * 70)

    # --------------------------------------------------------
    # Load data
    # --------------------------------------------------------

    if df is None:

        df = _load_day17_data()

    if not isinstance(
        df,
        pd.DataFrame
    ):

        raise TypeError(
            "run_day17(df) requires "
            "a pandas DataFrame."
        )

    if df.empty:

        raise ValueError(
            "Day 17 received an empty DataFrame."
        )

    print(
        f"Latest company records: {len(df)}"
    )

    # --------------------------------------------------------
    # Calculate composite score
    # --------------------------------------------------------

    scored_df = (
        calculate_composite_quality_score(
            df,
            sector_relative=True
        )
    )

    min_score = (
        scored_df[
            "composite_quality_score"
        ].min()
    )

    max_score = (
        scored_df[
            "composite_quality_score"
        ].max()
    )

    print()
    print(
        "Composite quality score calculated."
    )

    print(
        f"Score range: "
        f"{min_score:.2f} - "
        f"{max_score:.2f}"
    )

    # --------------------------------------------------------
    # Generate six preset results
    # --------------------------------------------------------

    preset_results = {}

    print()
    print("=" * 70)
    print("DAY 17 — PRESET RESULTS")
    print("=" * 70)

    for preset_name in DAY17_PRESETS:

        preset_df = (
            _apply_day17_preset_check(
                scored_df,
                preset_name
            )
        )

        preset_df = (
            preset_df
            .sort_values(
                "composite_quality_score",
                ascending=False,
                kind="mergesort"
            )
            .reset_index(drop=True)
        )

        preset_results[
            preset_name
        ] = preset_df

        print(
            f"{preset_name:<25}"
            f"{len(preset_df):>4} companies"
        )

    # --------------------------------------------------------
    # Export Excel
    # --------------------------------------------------------

    output_path = (
        export_screener_output(
            preset_results,
            SCREENER_OUTPUT_XLSX
        )
    )

    print()
    print("=" * 70)
    print("DAY 17 — EXPORT COMPLETE")
    print("=" * 70)

    print(
        f"Excel file: {output_path}"
    )

    print(
        f"Sheets created: "
        f"{len(preset_results)}"
    )

    for preset_name, result in (
        preset_results.items()
    ):

        print(
            f"  {preset_name}: "
            f"{len(result)} companies"
        )

    print()
    print(
        "SUCCESS: Day 17 composite score "
        "and screener export completed."
    )

    return (
        scored_df,
        preset_results
    )


# ============================================================
# END DAY 17
# ============================================================.