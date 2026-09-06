from pathlib import Path
import sqlite3
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# PATHS
# ============================================================

def table_exists(conn, table_name):
    """Return True when the requested SQLite table exists."""
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = ? AND name = ? LIMIT 1",
        ("table", table_name),
    ).fetchone()
    return row is not None



PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = PROJECT_ROOT / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_FILE = OUTPUT_DIR / "peer_comparison.xlsx"


# ============================================================
# REQUIRED METRICS
# ============================================================

METRICS = {
    "roe": "ROE",
    "roce": "ROCE",
    "npm": "NPM",
    "de": "D/E",
    "fcf": "FCF",
    "pat_cagr_5yr": "PAT CAGR 5yr",
    "revenue_cagr_5yr": "Revenue CAGR 5yr",
    "eps_cagr_5yr": "EPS CAGR 5yr",
    "interest_coverage": "Interest Coverage",
    "asset_turnover": "Asset Turnover",
}


# ============================================================
# LOAD PEER PERCENTILE DATA
# ============================================================

def load_peer_data():
    conn = sqlite3.connect(DB_PATH)

    query = """
        SELECT
            company_id,
            peer_group_name,
            metric,
            value,
            percentile_rank,
            year
        FROM peer_percentiles
        ORDER BY peer_group_name, company_id
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    return df



# ============================================================
# LOAD ALL PEER GROUPS
# ============================================================

EXPECTED_PEER_GROUPS = [
    "Automobiles",
    "Consumer Finance",
    "FMCG",
    "IT Services",
    "Life Insurance",
    "Oil & Gas",
    "Pharmaceuticals",
    "Power & Utilities",
    "Private Banks",
    "Public Sector Banks",
    "Steel",
]


def load_peer_groups():
    """
    Use peer_groups as the authoritative source for the 11 required
    sheets. Do not derive sheet names only from the latest
    peer_percentiles rows, because that can make valid groups disappear.
    """
    conn = sqlite3.connect(DB_PATH)

    try:
        columns = pd.read_sql_query(
            "PRAGMA table_info(peer_groups)",
            conn,
        )["name"].tolist()

        if "peer_group_name" not in columns:
            return EXPECTED_PEER_GROUPS.copy()

        rows = pd.read_sql_query(
            """
            SELECT DISTINCT peer_group_name
            FROM peer_groups
            WHERE peer_group_name IS NOT NULL
              AND TRIM(peer_group_name) <> ''
            ORDER BY peer_group_name
            """,
            conn,
        )
    finally:
        conn.close()

    groups = (
        rows["peer_group_name"]
        .astype(str)
        .str.strip()
        .tolist()
        if not rows.empty
        else []
    )

    for expected in EXPECTED_PEER_GROUPS:
        if expected not in groups:
            groups.append(expected)

    return sorted(set(groups))


# ============================================================
# LOAD COMPANY NAMES
# ============================================================

def load_company_names():
    conn = sqlite3.connect(DB_PATH)

    # First inspect available company columns.
    columns = pd.read_sql_query(
        "PRAGMA table_info(companies)",
        conn
    )["name"].tolist()

    # Try to find a ticker/company-id style column.
    possible_id_columns = [
        "id",
        "company_id",
        "symbol",
        "ticker",
        "nse_symbol",
        "stock_code",
        "code"
    ]

    id_column = next(
        (col for col in possible_id_columns if col in columns),
        None
    )

    if id_column:
        query = f"""
            SELECT
                {id_column} AS company_id,
                company_name
            FROM companies
        """

        names = pd.read_sql_query(query, conn)

    else:
        names = pd.DataFrame(
            columns=["company_id", "company_name"]
        )

    conn.close()

    return names


# ============================================================
# LOAD BENCHMARK COMPANIES
# ============================================================

def load_benchmarks():
    conn = sqlite3.connect(DB_PATH)

    query = """
        SELECT
            peer_group_name,
            company_id
        FROM peer_groups
        WHERE is_benchmark = 1
    """

    benchmarks = pd.read_sql_query(query, conn)

    conn.close()

    return benchmarks


# ============================================================
# BUILD WIDE TABLE
# ============================================================

def build_wide_table(group_df, company_names):
    """
    Convert long peer-percentile data into:
      company_id + company_name
      10 metric columns
      10 percentile columns

    The old implementation used:
        fillna(result.index.astype(str))
    which raises:
        TypeError: value parameter must be a scalar, dict or Series,
        but you passed a Index

    This version uses an index-aligned Series instead.
    """
    if group_df.empty:
        columns = (
            ["company_id", "company_name"]
            + list(METRICS.keys())
            + [
                f"{key}_percentile"
                for key in METRICS
            ]
        )
        return pd.DataFrame(columns=columns)

    group_df = group_df.copy()

    group_df["company_id"] = (
        group_df["company_id"]
        .astype(str)
        .str.strip()
    )

    # Keep the newest available record for each company + metric.
    group_df["_year_sort"] = (
        group_df["year"].astype(str)
    )

    group_df = (
        group_df
        .sort_values(
            ["company_id", "metric", "_year_sort"]
        )
        .drop_duplicates(
            subset=["company_id", "metric"],
            keep="last",
        )
        .drop(columns=["_year_sort"])
    )

    values = group_df.pivot_table(
        index="company_id",
        columns="metric",
        values="value",
        aggfunc="first",
    )

    percentiles = group_df.pivot_table(
        index="company_id",
        columns="metric",
        values="percentile_rank",
        aggfunc="first",
    )

    result = values.copy()

    # Add company names using company_id as the join key.
    if (
        company_names is not None
        and not company_names.empty
    ):
        names = (
            company_names[
                ["company_id", "company_name"]
            ]
            .drop_duplicates(
                subset=["company_id"],
                keep="first",
            )
            .set_index("company_id")
        )

        result = names.join(
            result,
            how="right",
        )

    # Safe fallback for missing company names.
    if "company_name" not in result.columns:
        result.insert(
            0,
            "company_name",
            pd.Series(
                result.index.astype(str),
                index=result.index,
            ),
        )
    else:
        fallback_names = pd.Series(
            result.index.astype(str),
            index=result.index,
        )

        result["company_name"] = (
            result["company_name"]
            .fillna(fallback_names)
            .astype(str)
        )

    # Ensure all 10 metric columns exist.
    for metric_key in METRICS:
        if metric_key not in result.columns:
            result[metric_key] = None

    # Add all 10 percentile columns.
    for metric_key in METRICS:
        percentile_column = (
            f"{metric_key}_percentile"
        )

        if metric_key in percentiles.columns:
            result[percentile_column] = (
                percentiles[metric_key]
            )
        else:
            result[percentile_column] = None

    result.index.name = "company_id"
    result = result.reset_index()

    result["company_id"] = (
        result["company_id"]
        .astype(str)
        .str.strip()
    )

    ordered_columns = (
        ["company_id", "company_name"]
        + list(METRICS.keys())
        + [
            f"{key}_percentile"
            for key in METRICS
        ]
    )

    return result[ordered_columns]


# ============================================================
# ADD MEDIAN ROW
# ============================================================

def add_median_row(df):
    median_values = {}

    for column in df.columns:
        if column in ["company_id", "company_name"]:
            continue

        numeric_values = pd.to_numeric(
            df[column],
            errors="coerce"
        )

        if numeric_values.notna().any():
            median_values[column] = numeric_values.median()
        else:
            median_values[column] = None

    median_row = {
        column: None
        for column in df.columns
    }

    median_row["company_id"] = ""
    median_row["company_name"] = "Peer Group Median"

    median_row.update(median_values)

    return pd.concat(
        [
            df,
            pd.DataFrame([median_row])
        ],
        ignore_index=True
    )


# ============================================================
# WRITE EXCEL FILE
# ============================================================

def generate_excel(
    peer_df,
    company_names,
    benchmarks,
    peer_groups,
):
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    print()
    print("=" * 60)
    print("PEER COMPARISON EXCEL REPORT")
    print("=" * 60)
    print(
        f"Peer groups found: {len(peer_groups)}"
    )

    # Remove the old workbook before creating a fresh one.
    if OUTPUT_FILE.exists():
        try:
            OUTPUT_FILE.unlink()
        except PermissionError as exc:
            raise PermissionError(
                "Cannot replace peer_comparison.xlsx. "
                "Close the Excel file if it is open, then run again."
            ) from exc

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        used_sheet_names = set()

        for peer_group in peer_groups:
            print(
                f"Creating sheet: {peer_group}"
            )

            group_df = peer_df[
                peer_df["peer_group_name"]
                == peer_group
            ].copy()

            wide_df = build_wide_table(
                group_df,
                company_names,
            )

            wide_df = add_median_row(
                wide_df
            )

            # Excel sheet names have a 31-character limit.
            sheet_name = str(
                peer_group
            ).strip()[:31]

            # Avoid duplicate sheet names after truncation.
            base_name = sheet_name or "Peer Group"
            sheet_name = base_name
            counter = 2

            while sheet_name in used_sheet_names:
                suffix = f" ({counter})"
                sheet_name = (
                    base_name[
                        :31 - len(suffix)
                    ]
                    + suffix
                )
                counter += 1

            used_sheet_names.add(sheet_name)

            wide_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

    print()
    print(
        f"Excel created: {OUTPUT_FILE}"
    )

    return OUTPUT_FILE


# ============================================================
# FORMAT EXCEL
# ============================================================

def format_excel(filename, benchmarks):
    wb = load_workbook(filename)

    # Required colours
    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C"
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    benchmark_fill = PatternFill(
        fill_type="solid",
        fgColor="FFD966"
    )

    median_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )

    header_font = Font(
        bold=True
    )

    # Benchmark lookup
    benchmark_lookup = {}

    if not benchmarks.empty:
        for _, row in benchmarks.iterrows():
            group = str(row["peer_group_name"])
            company_id = str(row["company_id"])

            benchmark_lookup.setdefault(
                group,
                set()
            ).add(company_id)

    for ws in wb.worksheets:

        # ----------------------------------------------------
        # HEADER
        # ----------------------------------------------------

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        ws.freeze_panes = "C2"

        # ----------------------------------------------------
        # COLUMN IDENTIFICATION
        # ----------------------------------------------------

        headers = {
            cell.column: str(cell.value)
            for cell in ws[1]
        }

        percentile_columns = [
            column
            for column, header in headers.items()
            if "percentile" in header.lower()
        ]

        # ----------------------------------------------------
        # PERCENTILE COLOURING
        # ----------------------------------------------------

        for row in range(2, ws.max_row + 1):

            # Last row is median row
            is_median = (
                ws.cell(row, 2).value
                == "Peer Group Median"
            )

            if is_median:
                continue

            for column in percentile_columns:

                cell = ws.cell(
                    row=row,
                    column=column
                )

                if cell.value is None:
                    continue

                try:
                    value = float(cell.value)
                except (TypeError, ValueError):
                    continue

                if value >= 0.75:
                    cell.fill = green_fill

                elif value >= 0.25:
                    cell.fill = yellow_fill

                else:
                    cell.fill = red_fill

        # ----------------------------------------------------
        # BENCHMARK ROW
        # ----------------------------------------------------

        sheet_group = ws.title

        benchmark_ids = set()

        for group_name, company_ids in benchmark_lookup.items():

            if str(group_name)[:31] == sheet_group:
                benchmark_ids = company_ids
                break

        for row in range(2, ws.max_row + 1):

            company_id = ws.cell(
                row=row,
                column=1
            ).value

            if company_id is None:
                continue

            if str(company_id) in benchmark_ids:

                for column in range(
                    1,
                    ws.max_column + 1
                ):
                    cell = ws.cell(
                        row=row,
                        column=column
                    )

                    cell.fill = benchmark_fill
                    cell.font = Font(
                        bold=True
                    )

        # ----------------------------------------------------
        # MEDIAN ROW
        # ----------------------------------------------------

        for row in range(
            2,
            ws.max_row + 1
        ):

            if (
                ws.cell(row, 2).value
                == "Peer Group Median"
            ):

                for column in range(
                    1,
                    ws.max_column + 1
                ):

                    cell = ws.cell(
                        row=row,
                        column=column
                    )

                    cell.fill = median_fill
                    cell.font = Font(
                        bold=True
                    )

        # ----------------------------------------------------
        # NUMBER FORMATTING
        # ----------------------------------------------------

        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row
        ):

            for cell in row:

                if cell.value is None:
                    continue

                header = headers.get(
                    cell.column,
                    ""
                )

                if "percentile" in header.lower():
                    cell.number_format = "0.0%"

                elif isinstance(
                    cell.value,
                    (int, float)
                ):
                    cell.number_format = "0.00"

        # ----------------------------------------------------
        # COLUMN WIDTHS
        # ----------------------------------------------------

        for column in ws.columns:

            max_length = 0

            for cell in column:

                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            column_letter = get_column_letter(
                column[0].column
            )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                28
            )

        ws.auto_filter.ref = ws.dimensions

    wb.save(filename)


# ============================================================
# VALIDATION
# ============================================================

def validate_output(filename):
    wb = load_workbook(
        filename,
        read_only=True
    )

    sheet_count = len(wb.sheetnames)

    print()
    print("=" * 60)
    print("VALIDATION")
    print("=" * 60)

    print(
        f"Sheets generated: {sheet_count}"
    )

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        print(
            f"  ✓ {sheet}: "
            f"{ws.max_row - 1} companies/data rows"
        )

    if sheet_count == 11:
        print("✓ Exactly 11 peer-group sheets")
    else:
        print(
            f"⚠ Expected 11 sheets, "
            f"found {sheet_count}"
        )

    wb.close()


# ============================================================
# MAIN
# ============================================================

def main():
    print()
    print(
        "Starting Day 20 - Peer Comparison Report"
    )

    print(f"\nProject root: {PROJECT_ROOT}")
    print(
        f"Database: {DB_PATH}"
    )
    print(
        f"Output: {OUTPUT_FILE}"
    )

    if not DB_PATH.exists():
        raise FileNotFoundError(
            f"Database not found: {DB_PATH}"
        )

    # --------------------------------------------------------
    # Load database data
    # --------------------------------------------------------

    peer_df = load_peer_data()

    if peer_df.empty:
        raise RuntimeError(
            "peer_percentiles table is empty or contains "
            "no supported metric records."
        )

    company_names = load_company_names()
    benchmarks = load_benchmarks()
    peer_groups = load_peer_groups()

    # --------------------------------------------------------
    # Database check
    # --------------------------------------------------------

    print()
    print("=" * 60)
    print("DATABASE CHECK")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)

    try:
        if table_exists(conn, "companies"):
            company_count = conn.execute(
                "SELECT COUNT(*) FROM companies"
            ).fetchone()[0]

            print(
                f"Companies in database: "
                f"{company_count}"
            )

        if table_exists(conn, "peer_groups"):
            group_rows = conn.execute(
                """
                SELECT
                    peer_group_name,
                    COUNT(DISTINCT company_id)
                FROM peer_groups
                WHERE peer_group_name IS NOT NULL
                GROUP BY peer_group_name
                ORDER BY peer_group_name
                """
            ).fetchall()

            print(
                f"Peer groups in database: "
                f"{len(group_rows)}"
            )

            for group_name, count in group_rows:
                print(
                    f"  {group_name}: "
                    f"{count} companies"
                )

    finally:
        conn.close()

    print(
        f"Peer percentile records: "
        f"{len(peer_df)}"
    )

    if not peer_df.empty:
        print(
            "Latest reporting year in table: "
            f"{peer_df['year'].astype(str).max()}"
        )

    print(
        f"Peer groups used for workbook: "
        f"{len(peer_groups)}"
    )

    print(
        f"Company names loaded: "
        f"{len(company_names)}"
    )

    print(
        f"Benchmark assignments: "
        f"{len(benchmarks)}"
    )

    # --------------------------------------------------------
    # Generate workbook
    # --------------------------------------------------------

    filename = generate_excel(
        peer_df,
        company_names,
        benchmarks,
        peer_groups,
    )

    # --------------------------------------------------------
    # Apply formatting
    # --------------------------------------------------------

    format_excel(
        filename,
        benchmarks,
    )

    # --------------------------------------------------------
    # Validate
    # --------------------------------------------------------

    validate_output(
        filename,
    )

    print()
    print("=" * 60)
    print("DAY 20 COMPLETE")
    print("=" * 60)
    print(
        f"Output: {filename}"
    )


if __name__ == "__main__":
    main()